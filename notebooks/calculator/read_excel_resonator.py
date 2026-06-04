from pathlib import Path
import argparse

import matplotlib.pyplot as plt
from openpyxl import load_workbook
import pandas as pd


DEFAULT_FOLDER = (
    r"d:\SynologyDrive\LiChiehHsiao\AS\SynologyDrive\simulation\chip design\EPR\resonator"
)
OUTPUT_FIGURE_NAME = "cpw_a_vs_qi.png"
OUTPUT_SUMS_FIGURE_NAME = "cpw_a_vs_calculation_sums.png"
OUTPUT_RATIO_FIGURE_NAME = "cpw_a_vs_ms_sa_ma_ratio.png"


def load_excel_files(folder: Path, recursive: bool = False) -> dict[str, dict[str, pd.DataFrame]]:
    """Load all Excel files in a folder and return nested dict[file][sheet] = DataFrame."""
    pattern = "**/*" if recursive else "*"
    candidates = [
        p
        for p in folder.glob(pattern)
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xls"}
    ]

    if not candidates:
        print(f"No Excel files found in: {folder}")
        return {}

    loaded: dict[str, dict[str, pd.DataFrame]] = {}
    for file_path in sorted(candidates):
        try:
            xls = pd.ExcelFile(file_path)
            sheets = {sheet: xls.parse(sheet_name=sheet) for sheet in xls.sheet_names}
            loaded[str(file_path)] = sheets
            print(f"Loaded: {file_path.name}")
            for sheet_name, df in sheets.items():
                print(f"  - {sheet_name}: {df.shape[0]} rows x {df.shape[1]} cols")
        except Exception as exc:
            print(f"Failed to read {file_path.name}: {exc}")

    return loaded


def _to_float(value: object) -> float:
    """Convert cell value to float, handling strings with commas/spaces."""
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        return float(cleaned)
    raise ValueError(f"Unsupported numeric value type: {type(value)}")


def extract_cpw_a_and_qi(file_path: Path) -> tuple[float, float]:
    """Read Summary!B5 (CPW a) and Summary!B19 (Q_i) from one Excel file."""
    summary = pd.read_excel(file_path, sheet_name="Summary", header=None)
    cpw_a = _to_float(summary.iat[4, 1])

    # Read cached formula result from Excel instead of formula text.
    wb = load_workbook(filename=file_path, data_only=True, read_only=True)
    try:
        ws = wb["Summary"]
        qi_raw = ws["B19"].value
    finally:
        wb.close()

    if qi_raw is None:
        raise ValueError("Summary!B19 has no cached value. Recalculate and save the workbook in Excel first.")

    qi = _to_float(qi_raw)
    return cpw_a, qi


def extract_cpw_a_and_calculation_sums(file_path: Path) -> tuple[float, float, float, float]:
    """Read Summary!B5 and grouped sums from Calculations using labels in column B."""
    wb = load_workbook(filename=file_path, data_only=True, read_only=True)
    try:
        ws_summary = wb["Summary"]
        ws_calc = wb["Calculations"]

        cpw_a_raw = ws_summary["B5"].value

        if cpw_a_raw is None:
            raise ValueError("Summary!B5 has no cached value. Recalculate and save the workbook in Excel first.")

        ms_sum = 0.0
        sa_sum = 0.0
        ma_sum = 0.0
        ms_found = False
        sa_found = False
        ma_found = False

        for row_idx, row in enumerate(
            ws_calc.iter_rows(min_row=5, max_col=6, values_only=True),
            start=5,
        ):
            label_raw = row[1] if len(row) > 1 else None
            value_raw = row[5] if len(row) > 5 else None
            if label_raw is None:
                continue

            label = str(label_raw).strip().upper()
            if label == "MS":
                if value_raw is None:
                    print(f"Warning {file_path.name} row {row_idx}: MS has empty column F; skipped")
                    continue
                try:
                    ms_sum += _to_float(value_raw)
                except ValueError:
                    print(
                        f"Warning {file_path.name} row {row_idx}: "
                        f"MS has non-numeric column F value {value_raw!r}; skipped"
                    )
                    continue
                ms_found = True
            elif label == "SA":
                if value_raw is None:
                    print(f"Warning {file_path.name} row {row_idx}: SA has empty column F; skipped")
                    continue
                try:
                    sa_sum += _to_float(value_raw)
                except ValueError:
                    print(
                        f"Warning {file_path.name} row {row_idx}: "
                        f"SA has non-numeric column F value {value_raw!r}; skipped"
                    )
                    continue
                sa_found = True
            elif label == "MA":
                if value_raw is None:
                    print(f"Warning {file_path.name} row {row_idx}: MA has empty column F; skipped")
                    continue
                try:
                    ma_sum += _to_float(value_raw)
                except ValueError:
                    print(
                        f"Warning {file_path.name} row {row_idx}: "
                        f"MA has non-numeric column F value {value_raw!r}; skipped"
                    )
                    continue
                ma_found = True
    finally:
        wb.close()

    cpw_a = _to_float(cpw_a_raw)
    if not ms_found or not sa_found or not ma_found:
        missing = [name for name, found in [("MS", ms_found), ("SA", sa_found), ("MA", ma_found)] if not found]
        raise ValueError(f"Missing category rows in Calculations column B: {', '.join(missing)}")

    return cpw_a, ms_sum, sa_sum, ma_sum


def plot_qi_vs_cpw_a(folder: Path, recursive: bool = False) -> Path | None:
    """Plot Q_i (y) as a function of CPW a (x) from all Excel files in folder."""
    pattern = "**/*" if recursive else "*"
    candidates = [
        p
        for p in folder.glob(pattern)
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xls"}
    ]

    if not candidates:
        print(f"No Excel files found in: {folder}")
        return None

    cpw_a_values: list[float] = []
    qi_values: list[float] = []
    used_files: list[str] = []

    for file_path in sorted(candidates):
        try:
            cpw_a, qi = extract_cpw_a_and_qi(file_path)
            cpw_a_values.append(cpw_a)
            qi_values.append(qi)
            used_files.append(file_path.name)
            print(f"Parsed {file_path.name}: CPW a={cpw_a}, Q_i={qi}")
        except Exception as exc:
            print(f"Skipped {file_path.name}: {exc}")

    if not cpw_a_values:
        print("No valid (CPW a, Q_i) data points were found.")
        return None

    ordered = sorted(zip(cpw_a_values, qi_values, used_files), key=lambda x: x[0])
    x_vals = [v[0] for v in ordered]
    y_vals = [v[1] for v in ordered]

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(x_vals, y_vals, marker="o", linestyle="-", linewidth=1.5)
    ax.set_xscale("log")
    ax.set_yscale("log")

    ax.set_xlabel("CPW a (um)")
    ax.set_ylabel("Q_i")
    ax.set_title("Q_i vs CPW a")
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    output_path = folder / OUTPUT_FIGURE_NAME
    fig.savefig(output_path, dpi=180)
    print(f"Saved figure: {output_path}")
    plt.show()

    return output_path


def plot_calculation_sums_vs_cpw_a(folder: Path, recursive: bool = False) -> Path | None:
    """Plot grouped sums from Calculations worksheet as functions of CPW a."""
    pattern = "**/*" if recursive else "*"
    candidates = [
        p
        for p in folder.glob(pattern)
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xls"}
    ]

    if not candidates:
        print(f"No Excel files found in: {folder}")
        return None

    cpw_a_values: list[float] = []
    sum_f9_f10_values: list[float] = []
    sum_f11_f12_values: list[float] = []
    sum_f13_f14_f15_values: list[float] = []
    used_files: list[str] = []

    for file_path in sorted(candidates):
        try:
            cpw_a, sum_f9_f10, sum_f11_f12, sum_f13_f14_f15 = extract_cpw_a_and_calculation_sums(file_path)
            cpw_a_values.append(cpw_a)
            sum_f9_f10_values.append(sum_f9_f10)
            sum_f11_f12_values.append(sum_f11_f12)
            sum_f13_f14_f15_values.append(sum_f13_f14_f15)
            used_files.append(file_path.name)
            print(
                f"Parsed {file_path.name}: CPW a={cpw_a}, "
                f"MS={sum_f9_f10}, SA={sum_f11_f12}, MA={sum_f13_f14_f15}"
            )
        except Exception as exc:
            print(f"Skipped {file_path.name} for calculation sums: {exc}")

    if not cpw_a_values:
        print("No valid calculation-sum data points were found.")
        return None

    ordered = sorted(
        zip(cpw_a_values, sum_f9_f10_values, sum_f11_f12_values, sum_f13_f14_f15_values, used_files),
        key=lambda x: x[0],
    )
    x_vals = [v[0] for v in ordered]
    y_f9_f10 = [v[1] for v in ordered]
    y_f11_f12 = [v[2] for v in ordered]
    y_f13_f14_f15 = [v[3] for v in ordered]

    fig, ax = plt.subplots(figsize=(9, 5))
    ax.plot(x_vals, y_f9_f10, marker="o", linestyle="-", linewidth=1.5, label="MS")
    ax.plot(x_vals, y_f11_f12, marker="s", linestyle="-", linewidth=1.5, label="SA")
    ax.plot(x_vals, y_f13_f14_f15, marker="^", linestyle="-", linewidth=1.5, label="MA")
    ax.set_xscale("log")
    ax.set_yscale("log")
    ax.set_xlabel("CPW a (um)")
    ax.set_ylabel("Value")
    ax.set_title("Calculations Sums vs CPW a")
    ax.grid(True, alpha=0.3)
    ax.legend()
    fig.tight_layout()

    output_path = folder / OUTPUT_SUMS_FIGURE_NAME
    fig.savefig(output_path, dpi=180)
    print(f"Saved figure: {output_path}")
    plt.show()

    return output_path


def plot_calculation_ratios_vs_cpw_a(folder: Path, recursive: bool = False) -> Path | None:
    """Plot MS/SA/MA normalized ratios as functions of CPW a."""
    pattern = "**/*" if recursive else "*"
    candidates = [
        p
        for p in folder.glob(pattern)
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xls"}
    ]

    if not candidates:
        print(f"No Excel files found in: {folder}")
        return None

    cpw_a_values: list[float] = []
    ms_ratio_values: list[float] = []
    sa_ratio_values: list[float] = []
    ma_ratio_values: list[float] = []
    used_files: list[str] = []

    for file_path in sorted(candidates):
        try:
            cpw_a, ms_sum, sa_sum, ma_sum = extract_cpw_a_and_calculation_sums(file_path)
            total = ms_sum + sa_sum + ma_sum
            if total == 0:
                print(f"Skipped {file_path.name} for ratio plot: MS+SA+MA is zero")
                continue

            cpw_a_values.append(cpw_a)
            ms_ratio_values.append(ms_sum / total)
            sa_ratio_values.append(sa_sum / total)
            ma_ratio_values.append(ma_sum / total)
            used_files.append(file_path.name)
            print(
                f"Parsed ratio {file_path.name}: CPW a={cpw_a}, "
                f"MS={ms_sum / total:.4f}, SA={sa_sum / total:.4f}, MA={ma_sum / total:.4f}"
            )
        except Exception as exc:
            print(f"Skipped {file_path.name} for ratio plot: {exc}")

    if not cpw_a_values:
        print("No valid ratio data points were found.")
        return None

    ordered = sorted(
        zip(cpw_a_values, ms_ratio_values, sa_ratio_values, ma_ratio_values, used_files),
        key=lambda x: x[0],
    )
    x_vals = [v[0] for v in ordered]
    y_ms_ratio = [v[1] for v in ordered]
    y_sa_ratio = [v[2] for v in ordered]
    y_ma_ratio = [v[3] for v in ordered]

    fig, ax = plt.subplots(figsize=(9, 5))
    ax.plot(x_vals, y_ms_ratio, marker="o", linestyle="-", linewidth=1.5, label="MS ratio")
    ax.plot(x_vals, y_sa_ratio, marker="s", linestyle="-", linewidth=1.5, label="SA ratio")
    ax.plot(x_vals, y_ma_ratio, marker="^", linestyle="-", linewidth=1.5, label="MA ratio")
    ax.set_xscale("log")
    ax.set_xlabel("CPW a (um)")
    ax.set_ylabel("Ratio")
    ax.set_ylim(0.0, 1.0)
    ax.set_title("MS / SA / MA Ratios vs CPW a")
    ax.grid(True, alpha=0.3)
    ax.legend()
    fig.tight_layout()

    output_path = folder / OUTPUT_RATIO_FIGURE_NAME
    fig.savefig(output_path, dpi=180)
    print(f"Saved figure: {output_path}")
    plt.show()

    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Read Excel files in the resonator folder")
    parser.add_argument(
        "--folder",
        type=str,
        default=DEFAULT_FOLDER,
        help="Folder containing Excel files",
    )
    parser.add_argument(
        "--recursive",
        action="store_true",
        help="Search subfolders recursively",
    )
    parser.add_argument(
        "--inspect",
        action="store_true",
        help="Only inspect all sheets and print summary table",
    )
    args = parser.parse_args()

    folder = Path(args.folder)
    if not folder.exists() or not folder.is_dir():
        raise FileNotFoundError(f"Folder not found or not a directory: {folder}")

    if args.inspect:
        _ = load_excel_files(folder=folder, recursive=args.recursive)
    else:
        _ = plot_qi_vs_cpw_a(folder=folder, recursive=args.recursive)
        _ = plot_calculation_sums_vs_cpw_a(folder=folder, recursive=args.recursive)
        _ = plot_calculation_ratios_vs_cpw_a(folder=folder, recursive=args.recursive)


if __name__ == "__main__":
    main()
